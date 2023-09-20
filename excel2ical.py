import datetime
import os
import sys
import uuid
from openpyxl import load_workbook
from icalendar import Calendar, Event
from tkinter import (Button, Frame, Label, PhotoImage, Tk, filedialog, messagebox)

WINDOW_TITLE = "Excel2ical v2.2          (buc @ hems.de)"
WINDOW_ICON = "excel2ics.ico"
APP_TITLE = "Excel2ical"
MASCOT_PIC = "cal.png"
FONT_DESCRIPTION = ("Helvetica", 10)
FONT_BUTTONS = ("Helvetica", 12)
FONT_CONVERT_BUTTON = ("Helvetica", 12, "bold")
FONT_COLOR_CONVERT_BUTTON = "white"
FONT_LABEL = ("Helvetica", 10, "italic")
BG_COLOR = "gray80"
APP_DESCRIPTION = "Dieses Tool wandelt die Excelterminliste in eine\nOutlook-importierbare ICS-Datei um."
ICS_FILE_LABEL = "Kein ICS-Dateiname bestimmt"
ICS_BUTTON_LABEL = "ICS-Datei bestimmen"
EXCEL_FILE_LABEL = "Keine Excel-Datei ausgewählt"
EXCEL_BUTTON_LABEL = "Exceldatei auswählen"
CONVERT_BUTTON_TEXT = "ICS erzeugen"
BG_COLOR_CONVERT_BUTTON = "#ee2724"

def excel_to_ics(excel_file_path, ics_file_path):
    # Open the Excel file
    wb = load_workbook(excel_file_path)
    
    # Get the first sheet
    sheet = wb.worksheets[0]
    
    # Create a new ICS calendar
    cal = Calendar()
    
    # Set the calendar properties
    cal.add('prodid', '-//HEMS SCHULJAHRESKALENDER//mxm.dk//')
    cal.add('version', '2.0')
    
    # Iterate over the rows in the sheet, starting from the second row, first row has the headlines
    #for row in sheet.iter_rows(min_row=2):
    for row_number, row in enumerate(sheet.iter_rows(min_row=4), start=4):
        # Create a new ICS event
        event = Event()
 
        event.add('uid', uuid.uuid4()) # UID of the event, each event needs a distinguished UID
        event.add('dtstamp', datetime.datetime.now()) # Date of creation of the event
        # Set the event properties using the data from the Excel file
        event.add('summary', row[6].value)
         
        start = row[0].value # start Date of the event

        if not all([cell.value is None for cell in row]): # for empty rows: skip
        
            # Check if information in start date is really a date
            if type(start) is not datetime.datetime:
                messagebox.showerror('Fehler', f'Eintrag "{start}" ist kein Startdatum.\nÜberspringe Zeile {row_number} in Excel...')
                continue

            # strip the information in the cell of the time (unnecessary and looks ugly in ICS)
            startd = start.date()
            
            # Combine date and time if time is given
            if type(row[2].value) == datetime.time:
                startd = datetime.datetime.combine(row[0].value, row[2].value)
            event.add('dtstart', startd)


            #now follows the enddate + endtime        
            if(row[3].value is not None):
                end = row[3].value # end date of the event
                
                # Check if information in end date is really a date
                if type(end) is not datetime.datetime:
                    messagebox.showerror('Fehler', f'Eintrag "{end}" ist kein Enddatum.\nÜberspringe Zeile {row_number} in Excel...')
                    continue
                
                # strip the information in the cell of the time (unnecessary)
                endd = end.date()
                               
                if type(row[5].value) == datetime.time: # check if end time is given
                    endd = datetime.datetime.combine(row[3].value, row[5].value)
                else:
                    endd = endd + datetime.timedelta(days=1) #needs to be done because otherwise it ends a day short

                event.add('dtend', endd)
          
            # if no enddate exists but an end time then the start date is used as a date plus the end time
            elif type(row[5].value) == datetime.time:
                    endd = datetime.datetime.combine(row[0].value, row[5].value)
                    event.add('dtend', endd)

                 
            if(row[6].value is not None):
                event.add('description', row[7].value)
            #if(row[7].value is not None):
            #    event.add('location', row[7].value)
            
            # Add the event to the calendar
            cal.add_component(event)

    # Write the calendar to the ICS file, exception handling
    try:
        with open(ics_file_path, 'wb') as f:
            f.write(cal.to_ical())
      
    except Exception as e:
        print(f"Fehler: {str(e)}")

def browse_excel_file():
    global exc_file_path
    exc_file_path = filedialog.askopenfilename(title='Excel-Datei zur Konvertierung auswählen', filetypes=[('Excel Dokument', '*.xlsx')]) 
    # next two lines to only show the filename on the label. The complete path would be too long to print.
    filename = os.path.basename(exc_file_path)
    excel_file_label.config(text=filename)
    
def browse_ics_file():
    global ics_file_path
    ics_file_path = filedialog.asksaveasfilename(title='ICS-Datei bestimmen', filetypes=[('ICS Dokument', '*.ics')], defaultextension='.ics') 
    # next two lines to only show the filename on the label. The complete path would be too long to print.
    filename = os.path.basename(ics_file_path)
    ics_file_label.config(text=filename)
    
def convert_files():
    if not exc_file_path or not ics_file_path:
        messagebox.showerror('Fehler', 'Bitte sowohl eine Excel-Datei auswählen,\nals auch den Namen einer ICS-Datei bestimmen')
    else:
        excel_to_ics(exc_file_path, ics_file_path)
        messagebox.showinfo('Erfolg', 'Die Datei ' + str(ics_file_path) + '\nwurde erfolgreich erzeugt')

#important for pyinstaller
def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

#create window
root = Tk()
root.iconbitmap(resource_path(WINDOW_ICON))
root.title(WINDOW_TITLE)

# Set the window not resizable
root.resizable(0, 0)

#create frames (for coloring background)
title_frame = Frame(root, bg=BG_COLOR)
title_frame.grid()
middle_frame = Frame(root)
middle_frame.grid()
bottom_frame = Frame(root, bg=BG_COLOR)
bottom_frame.grid()

#load image
pimage = PhotoImage(file=resource_path(MASCOT_PIC))
hems_logo = Label(title_frame, image=pimage, bg=BG_COLOR)
hems_logo.image = pimage

#create labels and buttons
description_label = Label(title_frame, text=APP_DESCRIPTION, font=FONT_DESCRIPTION, bg=BG_COLOR)
excel_file_label = Label(middle_frame, text=EXCEL_FILE_LABEL, font=FONT_LABEL)
browse_excel_button = Button(middle_frame, text=EXCEL_BUTTON_LABEL, command=browse_excel_file, font=FONT_BUTTONS)
ics_file_label = Label(middle_frame, text=ICS_FILE_LABEL, font=FONT_LABEL)
browse_ics_button = Button(middle_frame, text=ICS_BUTTON_LABEL, command=browse_ics_file, font=FONT_BUTTONS)
convert_button = Button(bottom_frame, text=CONVERT_BUTTON_TEXT, command=convert_files, font=FONT_CONVERT_BUTTON, bg=BG_COLOR_CONVERT_BUTTON)

#position labels, image and buttons
hems_logo.grid(row=0, column=0, padx=4, pady=10)
description_label.grid(row=0, column=1, pady=10)

excel_file_label.grid(row=2, column=0, columnspan=2)
browse_excel_button.grid(row=3, column=0, pady=10, columnspan=2)

ics_file_label.grid(row=4, column=0, columnspan=2)
browse_ics_button.grid(row=5, column=0, pady=10, columnspan=2)

convert_button.grid(row=7, column=0, padx=128, pady=10, columnspan=2)

root.mainloop()
